
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
from datetime import datetime  # Para formatear fecha y hora
import re
import os

import requests

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio
import openpyxl  # Para crear archivo Excel
from flask import send_file


def procesar_form_empleado(dataForm, foto_perfil):
    # Formateando Salario
    salario_sin_puntos = re.sub('[^0-9]+', '', dataForm['salario_empleado'])
    # convertir salario a INT
    salario_entero = int(salario_sin_puntos)

    result_foto_perfil = procesar_imagen_perfil(foto_perfil)
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_empleados (nombre_empleado, apellido_empleado, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado, foto_empleado, salario_empleado) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_empleado'], dataForm['apellido_empleado'], dataForm['sexo_empleado'],
                        dataForm['telefono_empleado'], dataForm['email_empleado'], dataForm['profesion_empleado'], result_foto_perfil, salario_entero)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_empleado: {str(e)}'


def procesar_imagen_perfil(foto):
    try:
        # Nombre original del archivo
        filename = secure_filename(foto.filename)
        extension = os.path.splitext(filename)[1]

        # Creando un string de 50 caracteres
        nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
        nombreFile = nuevoNameFile + extension

        # Construir la ruta completa de subida del archivo
        basepath = os.path.abspath(os.path.dirname(__file__))
        upload_dir = os.path.join(basepath, f'../static/fotos_empleados/')

        # Validar si existe la ruta y crearla si no existe
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            # Dando permiso a la carpeta
            os.chmod(upload_dir, 0o755)

        # Construir la ruta completa de subida del archivo
        upload_path = os.path.join(upload_dir, nombreFile)
        foto.save(upload_path)

        return nombreFile

    except Exception as e:
        print("Error al procesar archivo:", e)
        return []


# Lista de Empleados
def sql_lista_empleadosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.dni_empleado,
                        e.direccion_empleado,
                        e.fecha_nacimiento,
                        e.salario_empleado,
                        e.foto_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_empleadosBD: {e}")
        return None


# Detalles del Empleado
def sql_detalles_empleadosBD(idEmpleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.dni_empleado,
                        e.direccion_empleado,
                        e.fecha_nacimiento,
                        e.salario_empleado,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado,
                        e.telefono_empleado, 
                        e.email_empleado,
                        e.profesion_empleado,
                        e.foto_empleado,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_empleados AS e
                    WHERE id_empleado =%s
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL, (idEmpleado,))
                empleadosBD = cursor.fetchone()
        return empleadosBD
    except Exception as e:
        print(
            f"Errro en la función sql_detalles_empleadosBD: {e}")
        return None


# Funcion Empleados Informe (Reporte)
def empleadosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_empleado,
                        e.nombre_empleado, 
                        e.apellido_empleado,
                        e.dni_empleado,
                        e.direccion_empleado,
                        e.fecha_nacimiento,
                        e.salario_empleado,
                        e.email_empleado,
                        e.telefono_empleado,
                        e.profesion_empleado,
                        DATE_FORMAT(e.fecha_registro, '%d de %b %Y %h:%i %p') AS fecha_registro,
                        CASE
                            WHEN e.sexo_empleado = 1 THEN 'Masculino'
                            ELSE 'Femenino'
                        END AS sexo_empleado
                    FROM tbl_empleados AS e
                    ORDER BY e.id_empleado DESC
                    """)
                cursor.execute(querySQL,)
                empleadosBD = cursor.fetchall()
        return empleadosBD
    except Exception as e:
        print(
            f"Error en la función empleadosReporte: {e}")
        return None


def generarReporteExcel():
    dataEmpleados = empleadosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Apellido", "DNI", "Dirección", "Fecha de Nacimiento","Sexo",
                     "Telefono", "Email", "Profesión", "Salario", "Fecha de Ingreso")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_peruana = '"S/" #,##0'

    # Agregar los registros a la hoja
    for registro in dataEmpleados:
        nombre_empleado = registro['nombre_empleado']
        apellido_empleado = registro['apellido_empleado']
        dni_empleado = registro['dni_empleado']
        direccion_empleado = registro['direccion_empleado']
        fecha_nacimiento = registro['fecha_nacimiento']
        sexo_empleado = registro['sexo_empleado']
        telefono_empleado = registro['telefono_empleado']
        email_empleado = registro['email_empleado']
        profesion_empleado = registro['profesion_empleado']
        salario_empleado = registro['salario_empleado']
        fecha_registro = registro['fecha_registro']

        # Agregar los valores a la hoja
        hoja.append((nombre_empleado, apellido_empleado, dni_empleado, direccion_empleado, fecha_nacimiento, sexo_empleado, telefono_empleado, email_empleado, profesion_empleado,
                     salario_empleado, fecha_registro))

    # Itera a través de las filas y aplica el formato a la columna de Salario
    # Se movió fuera del bucle principal para mayor eficiencia
    for fila_num in range(2, hoja.max_row + 1):
        columna = 10  # Columna J para Salario
        celda = hoja.cell(row=fila_num, column=columna)
        celda.number_format = formato_moneda_peruana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_empleados_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarEmpleadoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.salario_empleado,
                            CASE
                                WHEN e.sexo_empleado = 1 THEN 'Masculino'
                                ELSE 'Femenino'
                            END AS sexo_empleado
                        FROM tbl_empleados AS e
                        WHERE e.nombre_empleado LIKE %s 
                        ORDER BY e.id_empleado DESC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoBD: {e}")
        return []


def buscarEmpleadoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_empleado,
                            e.nombre_empleado, 
                            e.apellido_empleado,
                            e.sexo_empleado,
                            e.telefono_empleado,
                            e.email_empleado,
                            e.profesion_empleado,
                            e.salario_empleado,
                            e.foto_empleado
                        FROM tbl_empleados AS e
                        WHERE e.id_empleado =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                empleado = mycursor.fetchone()
                return empleado

    except Exception as e:
        print(f"Ocurrió un error en def buscarEmpleadoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                # Extraer y procesar datos del formulario
                nombre_empleado = data.form['nombre_empleado']
                apellido_empleado = data.form['apellido_empleado']
                sexo_empleado = data.form['sexo_empleado']
                telefono_empleado = data.form['telefono_empleado']
                email_empleado = data.form['email_empleado']
                profesion_empleado = data.form['profesion_empleado']

                # Procesar salario eliminando caracteres no numéricos
                salario_sin_puntos = re.sub('[^0-9]+', '', data.form['salario_empleado'])
                salario_empleado = int(salario_sin_puntos)
                id_empleado = data.form['id_empleado']

                # Construir consulta SQL y parámetros dinámicamente
                query_base = """
                    UPDATE tbl_empleados
                    SET 
                        nombre_empleado = %s,
                        apellido_empleado = %s,
                        sexo_empleado = %s,
                        telefono_empleado = %s,
                        email_empleado = %s,
                        profesion_empleado = %s,
                        salario_empleado = %s
                """
                params = [
                    nombre_empleado, apellido_empleado, sexo_empleado,
                    telefono_empleado, email_empleado, profesion_empleado, salario_empleado
                ]

                # Verificar si se subió un archivo de foto
                if 'foto_empleado' in data.files and data.files['foto_empleado'].filename != '':
                    file = data.files['foto_empleado']
                    fotoForm = procesar_imagen_perfil(file)
                    query_base += ", foto_empleado = %s"
                    params.append(fotoForm)

                # Agregar condición WHERE
                query_base += " WHERE id_empleado = %s"
                params.append(id_empleado)

                # Ejecutar la consulta
                cursor.execute(query_base, params)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None


# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar uEmpleado
def eliminarEmpleado(id_empleado, foto_empleado):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_empleados WHERE id_empleado=%s"
                cursor.execute(querySQL, (id_empleado,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

                if resultado_eliminar:
                    # Eliminadon foto_empleado desde el directorio
                    basepath = path.dirname(__file__)
                    url_File = path.join(
                        basepath, '../static/fotos_empleados', foto_empleado)

                    if path.exists(url_File):
                        remove(url_File)  # Borrar foto desde la carpeta

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarEmpleado : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []


# consultas DNI con api 
def consultar_dni_api(dni):
    """Función corregida con mapeo de campos correcto"""
    try:
        if not dni or len(dni) != 8 or not dni.isdigit():
            return {
                "success": False,
                "message": "DNI debe tener exactamente 8 dígitos numéricos"
            }
        
        api_url = "https://api.consultasperu.com/api/v1/query"

        headers = {
            'Content-Type': 'application/json',
        }

        # FORMATO CORRECTO - con type_document
        data = {
            "token": "09462861dd51641fa8d9947793ae1156a27df903d1468230ec496a4eac628c84",
            "type_document": "dni",
            "document_number": dni
        }

        print(f"🔍 Consultando DNI: {dni}")
        print(f"🔍 Data enviada: {data}")

        response = requests.post(api_url, headers=headers, json=data, timeout=10)
        print(f"🔍 Respuesta API : {response.status_code}")

        if response.status_code == 200:
            result = response.json()
            print(f"✅ Datos recibidos: {result}")

            if result.get('success') and result.get('data'):
                persona = result['data']
                
                # MAPEO CORRECTO DE CAMPOS
                # La API devuelve:
                # - "name" (no "nombres")
                # - "first_last_name" (no "apellido_paterno") 
                # - "second_last_name" (no "apellido_materno")
                # - "date_of_birth" (no "fechaNacimiento")
                
                nombres = persona.get('name', '').strip().title()
                apellido_paterno = persona.get('first_last_name', '').strip().title()
                apellido_materno = persona.get('second_last_name', '').strip().title()
                fecha_nacimiento = persona.get('date_of_birth', '')
                
                # Si hay surname completo, usarlo como fallback
                if not apellido_paterno and not apellido_materno:
                    surname_completo = persona.get('surname', '').strip().title()
                    if surname_completo:
                        apellidos_split = surname_completo.split(' ', 1)
                        apellido_paterno = apellidos_split[0] if len(apellidos_split) > 0 else ''
                        apellido_materno = apellidos_split[1] if len(apellidos_split) > 1 else ''
                
                # Construir nombre completo
                nombre_completo = f"{nombres} {apellido_paterno} {apellido_materno}".strip()
                
                return {
                    "success": True,
                    "data": {
                        "dni": persona.get('number', dni),
                        "nombres": nombres,
                        "apellido_paterno": apellido_paterno,
                        "apellido_materno": apellido_materno,
                        "apellidos": f"{apellido_paterno} {apellido_materno}".strip(),
                        "nombre_completo": nombre_completo,
                        "fecha_nacimiento": procesar_fecha_nacimiento(fecha_nacimiento),
                        "genero": persona.get('gender', ''),
                        "departamento": persona.get('department', ''),
                        "provincia": persona.get('province', ''),
                        "distrito": persona.get('district', ''),
                        "direccion": persona.get('address', ''),
                        "ubigeo": persona.get('ubigeo', ''),
                        # Campos adicionales útiles
                        "nombre_completo_api": persona.get('full_name', '').strip().title(),
                        "codigo_verificacion": persona.get('verification_code', ''),
                        "raw_data": persona  # Para debug y campos adicionales
                    }
                }
            else:
                return {
                    "success": False,
                    "message": "No se encontraron datos para este DNI"
                }
        elif response.status_code == 404:
            return {
                    "success": False,
                    "message": "DNI no encontrado en la base de datos"
            }
        elif response.status_code == 429:
            return {
                    "success": False,
                    "message": "Límite de consultas excedido. Intente más tarde"
            }
        else:
            print(f"❌ Error API: {response.text}")
            return {
                "success": False,
                "message": f"Error en la consulta: {response.status_code}"
            }
            
    except requests.exceptions.Timeout:
        return {
            "success": False,
            "message": "Tiempo de espera agotado. Intente nuevamente"
        }
    except requests.exceptions.ConnectionError:
        return {
            "success": False,
            "message": "Error de conexión con el servidor"
        }
    except requests.exceptions.RequestException as e:
        print(f"Error en la consulta: {e}")
        return {
            "success": False,
            "message": "Error al consultar la API externa"
        }
    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        return {
            "success": False,
            "message": "Error interno del servidor"
        }
    
def procesar_fecha_nacimiento(fecha_str):
    if not fecha_str:
        return None
    try:
        formatos= [
            "%Y-%m-%d",  # Formato YYYY-MM-DD
            "%d/%m/%Y",  # Formato DD/MM/YYYY
            "%d-%m-%Y"   # Formato DD-MM-YYYY
        ]

        for formato in formatos:
            try:
                fecha_obj = datetime.strptime(str(fecha_str), formato)
                return fecha_obj.strftime("%Y-%m-%d")  # Formato YYYY/MM/DD
            except ValueError:
                continue
        print(f"⚠️ Formato de fecha no reconocido: {fecha_str}") 
        return None
    except Exception as e:
        print(f"Error procesando fecha: {e}")
        return None
    
 # Función para configurar la API (agregar a tu configuración)
def configurar_api_dni():
    return {
        'api_url':'https://api.consultasperu.com/api/v1/query',
        'token': '09462861dd51641fa8d9947793ae1156a27df903d1468230ec496a4eac628c84',
        'timeout': 10,
        'activa': True,  # Cambia a False para desactivar la API
    }
